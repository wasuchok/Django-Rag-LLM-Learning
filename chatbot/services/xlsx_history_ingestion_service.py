from __future__ import annotations

import hashlib
from datetime import date, datetime
from pathlib import Path
from typing import Any

from django.db import transaction

from ..models import KnowledgeDocument
from .knowledge_access_service import get_knowledge_visibility_label
from .rag_service import delete_document_from_index, index_document

HISTORY_WORKBOOK_SHEET_NAME = "History-2024"
HISTORY_WORKBOOK_SHEET_PREFIX = "history-"
PREVIEW_DOCUMENT_LIMIT = 8
REQUIRED_HISTORY_COLUMNS = {"date", "machine_no", "problem", "action"}


def _load_openpyxl():
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise ValueError(
            "ยังไม่รองรับไฟล์ XLSX เพราะยังไม่ได้ติดตั้งแพ็กเกจ openpyxl"
        ) from exc

    return load_workbook


def _normalize_cell_text(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    if text in {"-", "[NULL]", "NULL", "None"}:
        return ""

    return " ".join(text.split())


def _normalize_header_text(value: Any) -> str:
    text = _normalize_cell_text(value).lower()
    return " ".join(text.split())


def _detect_history_column_key(value: Any) -> str | None:
    normalized = _normalize_header_text(value)
    raw_text = _normalize_cell_text(value)

    if not normalized:
        return None

    if normalized.startswith("item"):
        return "item"

    if "date" in normalized and "diff" not in normalized and "finish" not in normalized:
        return "date"

    if "machine no" in normalized or "machine no." in normalized or "เครื่องจักร" in raw_text:
        return "machine_no"

    if normalized.startswith("section"):
        return "section"

    if "bm" in normalized and "upm" in normalized and "others" in normalized:
        return "maintenance_type"

    if normalized == "cause":
        return "cause"

    if "problem" in normalized:
        return "problem"

    if normalized.startswith("action"):
        return "action"

    if "sub code" in normalized:
        return "sub_code"

    if "กำหนดผู้รับผิดชอบ" in raw_text or "mt,pdt" in normalized:
        return "assignee"

    if "repair by" in normalized or "ซ่อมโดย" in raw_text:
        return "repair_by"

    if "loss time" in normalized:
        return "loss_time"

    if "ค่าใช้จ่าย" in raw_text:
        return "cost"

    return None


def _find_history_header_row(worksheet) -> tuple[int, dict[int, str]]:
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=10, values_only=True),
        start=1,
    ):
        column_map: dict[int, str] = {}

        for column_index, value in enumerate(row, start=1):
            key = _detect_history_column_key(value)
            if key and key not in column_map.values():
                column_map[column_index] = key

        if REQUIRED_HISTORY_COLUMNS.issubset(set(column_map.values())):
            return row_index, column_map

    raise ValueError(
        f"ไม่พบหัวตารางที่รองรับในชีต {worksheet.title}"
    )


def _normalize_sheet_name(value: str) -> str:
    return " ".join(value.strip().lower().split())


def _resolve_history_sheet_name(workbook, requested_sheet_name: str | None) -> str:
    normalized_sheet_names = {
        _normalize_sheet_name(sheet_name): sheet_name
        for sheet_name in workbook.sheetnames
    }

    if requested_sheet_name:
        normalized_requested = _normalize_sheet_name(requested_sheet_name)
        matched_sheet_name = normalized_sheet_names.get(normalized_requested)
        if matched_sheet_name:
            return matched_sheet_name

        available_sheets = ", ".join(workbook.sheetnames)
        raise ValueError(
            f"ไม่พบชีต {requested_sheet_name} ในไฟล์นี้ ชีตที่มีคือ: {available_sheets}"
        )

    preferred_candidates = [
        sheet_name
        for sheet_name in workbook.sheetnames
        if _normalize_sheet_name(sheet_name).startswith(HISTORY_WORKBOOK_SHEET_PREFIX)
    ]
    if preferred_candidates:
        return preferred_candidates[0]

    detected_candidates: list[str] = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        try:
            _find_history_header_row(worksheet)
        except ValueError:
            continue
        detected_candidates.append(sheet_name)

    if len(detected_candidates) == 1:
        return detected_candidates[0]

    if len(detected_candidates) > 1:
        candidate_text = ", ".join(detected_candidates)
        raise ValueError(
            "พบหลายชีตที่มีโครงสร้างคล้ายประวัติซ่อม กรุณาระบุชื่อชีตด้วย --sheet: "
            f"{candidate_text}"
        )

    available_sheets = ", ".join(workbook.sheetnames)
    raise ValueError(
        "ไม่พบชีตที่รองรับสำหรับ import ประวัติซ่อมในไฟล์นี้ "
        f"ชีตที่มีคือ: {available_sheets}"
    )


def _is_meaningful_history_row(row_data: dict[str, str]) -> bool:
    if not any(row_data.values()):
        return False

    essential_values = [
        row_data.get("machine_no", ""),
        row_data.get("problem", ""),
        row_data.get("action", ""),
        row_data.get("cause", ""),
    ]
    return any(value.strip() for value in essential_values)


def extract_history_rows_from_xlsx(
    file_path: Path,
    *,
    sheet_name: str | None = None,
) -> list[dict[str, str]]:
    load_workbook = _load_openpyxl()
    workbook = load_workbook(
        filename=file_path,
        read_only=True,
        data_only=True,
    )

    try:
        resolved_sheet_name = _resolve_history_sheet_name(workbook, sheet_name)
        worksheet = workbook[resolved_sheet_name]
        header_row_index, column_map = _find_history_header_row(worksheet)
        rows: list[dict[str, str]] = []

        for row_index, row in enumerate(
            worksheet.iter_rows(
                min_row=header_row_index + 2,
                values_only=True,
            ),
            start=header_row_index + 2,
        ):
            row_data: dict[str, str] = {
                "sheet_name": resolved_sheet_name,
                "row_number": str(row_index),
            }

            for column_index, key in column_map.items():
                if column_index - 1 >= len(row):
                    row_data[key] = ""
                    continue

                row_data[key] = _normalize_cell_text(row[column_index - 1])

            if _is_meaningful_history_row(row_data):
                rows.append(row_data)

        return rows
    finally:
        workbook.close()


def build_history_row_title(row_data: dict[str, str]) -> str:
    parts = [
        row_data.get("date", ""),
        row_data.get("machine_no", ""),
        row_data.get("problem", ""),
    ]
    title = " | ".join(part for part in parts if part).strip()
    return title[:255] if len(title) > 255 else (title or "Machine History Case")


def build_history_row_content(
    row_data: dict[str, str],
    *,
    file_name: str,
) -> str:
    sections = [
        f"ต้นทางไฟล์: {file_name}",
        f"ชีต: {row_data.get('sheet_name', HISTORY_WORKBOOK_SHEET_NAME)}",
        f"แถวในชีต: {row_data.get('row_number', '-')}",
    ]

    field_mapping = [
        ("item", "ลำดับ"),
        ("date", "วันที่"),
        ("machine_no", "เครื่องจักร"),
        ("section", "แผนก"),
        ("maintenance_type", "ประเภทงาน"),
        ("cause", "สาเหตุ"),
        ("problem", "อาการ"),
        ("action", "การแก้ไข"),
        ("sub_code", "รหัสย่อย"),
        ("assignee", "ผู้รับผิดชอบ"),
        ("repair_by", "ผู้ปฏิบัติงาน"),
        ("loss_time", "เวลาสูญเสีย"),
        ("cost", "ค่าใช้จ่าย"),
    ]

    for key, label in field_mapping:
        value = row_data.get(key, "")
        if value:
            sections.append(f"{label}: {value}")

    return "\n".join(sections).strip()


def build_history_row_source(row_data: dict[str, str]) -> str:
    sheet_name = row_data.get("sheet_name", HISTORY_WORKBOOK_SHEET_NAME)
    key_parts = [
        sheet_name,
        row_data.get("item", ""),
        row_data.get("date", ""),
        row_data.get("machine_no", ""),
        row_data.get("problem", ""),
        row_data.get("row_number", ""),
    ]
    digest = hashlib.sha1("|".join(key_parts).encode("utf-8")).hexdigest()[:16]
    return f"xlsx-history:{sheet_name}:{digest}"


def _upsert_history_document(
    *,
    title: str,
    content: str,
    source: str,
    user_id: int | None,
    visibility: str,
) -> dict[str, Any]:
    existing_document = (
        KnowledgeDocument.objects.filter(source=source)
        .order_by("id")
        .first()
    )

    if existing_document is None:
        with transaction.atomic():
            document = KnowledgeDocument.objects.create(
                owner_id=user_id,
                title=title,
                content=content,
                source=source,
                visibility=visibility,
            )
            index_document(document)

        return {
            "document_id": document.id,
            "title": document.title,
            "source": document.source,
            "characters": len(content),
            "visibility": document.visibility,
            "visibility_label": get_knowledge_visibility_label(document.visibility),
            "status": "created",
        }

    if (
        existing_document.title == title
        and existing_document.content == content
        and existing_document.visibility == visibility
        and existing_document.owner_id == user_id
    ):
        return {
            "document_id": existing_document.id,
            "title": existing_document.title,
            "source": existing_document.source,
            "characters": len(existing_document.content or ""),
            "visibility": existing_document.visibility,
            "visibility_label": get_knowledge_visibility_label(existing_document.visibility),
            "status": "skipped",
        }

    with transaction.atomic():
        delete_document_from_index(existing_document.id)
        existing_document.owner_id = user_id
        existing_document.title = title
        existing_document.content = content
        existing_document.source = source
        existing_document.visibility = visibility
        existing_document.save(
            update_fields=["owner", "title", "content", "source", "visibility"]
        )
        index_document(existing_document)

    return {
        "document_id": existing_document.id,
        "title": existing_document.title,
        "source": existing_document.source,
        "characters": len(content),
        "visibility": existing_document.visibility,
        "visibility_label": get_knowledge_visibility_label(existing_document.visibility),
        "status": "updated",
    }


def summarize_file_ingestion(
    *,
    file_name: str,
    documents: list[dict[str, Any]],
    mode: str,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    created_count = sum(1 for item in documents if item.get("status") == "created")
    updated_count = sum(1 for item in documents if item.get("status") == "updated")
    skipped_count = sum(1 for item in documents if item.get("status") == "skipped")

    return {
        "file_name": file_name,
        "mode": mode,
        "sheet_name": sheet_name,
        "document_count": len(documents),
        "created_count": created_count,
        "updated_count": updated_count,
        "skipped_count": skipped_count,
        "documents_preview": documents[:PREVIEW_DOCUMENT_LIMIT],
        "documents": documents,
    }


def ingest_history_workbook(
    file_path: str | Path,
    *,
    display_name: str | None = None,
    user_id: int | None = None,
    visibility: str = KnowledgeDocument.VISIBILITY_SHARED,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    path = Path(file_path)
    source_file_name = display_name or path.name
    rows = extract_history_rows_from_xlsx(path, sheet_name=sheet_name)

    if not rows:
        raise ValueError(
            f"ไม่พบแถวข้อมูลที่ใช้งานได้ในชีต {sheet_name or 'auto-detect'}"
        )

    documents: list[dict[str, Any]] = []
    for row_data in rows:
        title = build_history_row_title(row_data)
        content = build_history_row_content(
            row_data,
            file_name=source_file_name,
        )
        source = build_history_row_source(row_data)
        documents.append(
            _upsert_history_document(
                title=title,
                content=content,
                source=source,
                user_id=user_id,
                visibility=visibility,
            )
        )

    return summarize_file_ingestion(
        file_name=source_file_name,
        documents=documents,
        mode="xlsx_history_rows",
        sheet_name=rows[0].get("sheet_name"),
    )
